import io
import os
import tempfile
import time
import uuid
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
import openpyxl
import win32com.client

from .models import AttributeGroup, CoopAttribute, CoopAttributeValue, CoopDeleteRequest, Cutting_factory, CuttingAround, CuttingSaw, PriceAttribute
from django.http import HttpResponseForbidden
from StoneFlow.models import coops
from mines.models import Mine
from users.models import Buyer, Profile, Warehouse, mother_material , raw_material
# Create your views here.
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from decimal import Decimal
from django.core.files.base import ContentFile
import base64
from django.db.models import Q
from django.conf import settings

from .models import Driver, coops, STATE_CHOICES
# views.py
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.template.loader import render_to_string

from openpyxl import Workbook
from django.http import HttpResponse

import jdatetime
from datetime import datetime, timezone
import qrcode
from io import BytesIO
import base64
from django.contrib.auth.decorators import user_passes_test


from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt



def get_allowed_confirm_users(stepNumber:int,user):

    access = StepAccess.objects.filter(step=stepNumber,user=user)

    allowed_roles = []

    if access.exists():
            access = access.first()
            if access.access_level=='submit':
                return True
    return False
    if stepNumber==1:
        allowed_roles = ['manager', 'fishzan','Programmer','Driver']  # Adjust based on your logic
        return allowed_roles

    if stepNumber==2:
        allowed_roles = ['manager', 'fishzan','Programmer']  # Adjust based on your logic
        return allowed_roles

    if stepNumber==3:
        allowed_roles = ['manager', 'fishzan','Programmer']  # Adjust based on your logic
        return allowed_roles
  
    if stepNumber==4:
        allowed_roles = ['manager', 'fishzan']  # Adjust based on your logic
        return allowed_roles  




def get_submit_and_confirmed(user,stepNumber):
        try:
            # user_profile = Profile.objects.get(user=user)
            # user_profile = User.objects.get(user=user)
            # user_role = user_profile.job_position.name
            can_submit = get_allowed_confirm_users(stepNumber=stepNumber,user=user)
            # Check if the user has access to submit this step
            # can_submit = user_role in allowed_roles
            # is_confirmed = check_order_confirmed(order=ret,stepNumber=1)
            is_confirmed = False
            return can_submit , is_confirmed
        except:
            return False,False






def coops_by_state(request):
    selected_state = request.GET.get('state')
    selected_material_id = request.GET.get('material')
    page = int(request.GET.get('page', 1))

    coop_list = coops.objects.filter(is_active=True).order_by('-id')
    steps = Step.objects.order_by('order')  # Ù…Ø±ØªØ¨â€ŒØ³Ø§Ø²ÛŒ Ù…Ø±Ø§Ø­Ù„
    materials = raw_material.objects.all()

    if selected_state:
        coop_list = coop_list.filter(state=selected_state)
    if selected_material_id:
        coop_list = coop_list.filter(material_id=selected_material_id)

    paginator = Paginator(coop_list, 10)
    page_obj = paginator.get_page(page)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('partials/coop_cards.html', {'coops': page_obj})
        return JsonResponse({'html': html, 'has_next': page_obj.has_next(), 'next_page': page + 1})

    context = {
        'coops': page_obj,
        'steps': steps,
        'selected_state': selected_state,
        'materials': materials,
        'selected_material_id': selected_material_id,
    }
    return render(request, 'coop_list.html', context)


@login_required
def coop_detail(request, coop_id):
    coop = get_object_or_404(coops, id=coop_id)
    user_access_qs = StepAccess.objects.filter(user=request.user)
    step_access = {access.step.id: access.access_level for access in user_access_qs}

    totalPrice = calculate_total_price(coop=coop)
    coop.total_price = totalPrice
    short_card_group = AttributeGroup.objects.get(name='short_card')
    short_attributes = short_card_group.attributes.all()

    coop_values = {}
    values = CoopAttributeValue.objects.filter(coop=coop, attribute__in=short_attributes)
    coop_values[coop.id] = {val.attribute.id: val.value for val in values}

    qr_data_url = generate_qr_code(request=request)

    return render(request, 'coop_detail.html', {'coop': coop,
                                                'step_access':step_access,
                                                'short_attributes': short_attributes,
                                                'coop_values': coop_values,
                                                'qr_data_url': qr_data_url,
                                                })



def generate_qr_code(request):
    try:
        current_url = request.build_absolute_uri()

        # Generate QR code image
        qr = qrcode.make(current_url)
        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        qr_data_url = f"data:image/png;base64,{qr_base64}"
    except:
        return None

    return qr_data_url




def coop_dashboard(request):
    material_id = request.GET.get('material_id')
    materials = raw_material.objects.all()
    selected_material = None
    qs = coops.objects.all()

    if material_id:
        qs = qs.filter(material_id=material_id)
        selected_material = raw_material.objects.filter(id=material_id).first()

    # # Ø¢Ù…Ø§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø¯Ø§Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ú†Ø§Ø±Øªâ€ŒÙ‡Ø§
    # state_counts = {}
    # for state_code, state_name in coops._meta.get_field('state').choices:
    #     state_counts[state_name] = qs.filter(state=state_code).count()


    from django.db.models import Count

    # Ø´Ù…Ø§Ø±Ø´ ØªØ¹Ø¯Ø§Ø¯ Ú©ÙˆÙ¾â€ŒÙ‡Ø§ Ø¯Ø± Ù‡Ø± ÙˆØ¶Ø¹ÛŒØª (Ù…Ø±Ø­Ù„Ù‡)
    state_counts_raw = qs.values('state__title').annotate(count=Count('id'))

    # ØªØ¨Ø¯ÛŒÙ„ Ø¨Ù‡ Ø¯ÛŒÚ©Ø´Ù†Ø±ÛŒ Ù‚Ø§Ø¨Ù„ Ø§Ø³ØªÙØ§Ø¯Ù‡: { 'Ø¹Ù†ÙˆØ§Ù† ÙˆØ¶Ø¹ÛŒØª': ØªØ¹Ø¯Ø§Ø¯ }
    state_counts = {
        item['state__title']: item['count']
        for item in state_counts_raw
    }



    chart_labels = list(state_counts.keys())
    chart_data = list(state_counts.values())

    return render(request, 'coop_dashboard.html', {
        'materials': materials,
        'selected_material': selected_material,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    })


from django.db import transaction
@login_required
@transaction.atomic
def create_coope(request):
    stepNumber = 1


    if request.method == 'POST':
        coop_record = None  # Ø¯Ø± Ø³Ø·Ø­ Ø¨Ø§Ù„Ø§ ØªØ¹Ø±ÛŒÙ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ… Ú©Ù‡ Ø¯Ø± except Ù‡Ù… Ù‚Ø§Ø¨Ù„ Ø¯Ø³ØªØ±Ø³ÛŒ Ø¨Ø§Ø´Ø¯
    # try:
        data = dict(request.POST.dict())
        data.pop('csrfmiddlewaretoken', None)

        image = None
        image_data = request.POST.get('image_data')
        if image_data:
            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            image = ContentFile(base64.b64decode(imgstr), name='captured_image.' + ext)
            data.pop('image_data', None)

        mine_id = request.POST.get('mine_id')
        if not mine_id:
            messages.error(request, "Ù„Ø·ÙØ§Ù‹ ÛŒÚ© Ù…Ø¹Ø¯Ù† Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯.")
            raise Exception("Ù…Ø¹Ø¯Ù† Ø§Ù†ØªØ®Ø§Ø¨ Ù†Ø´Ø¯Ù‡")

        selected_mine = Mine.objects.filter(id=mine_id).first()
        if not selected_mine:
            messages.error(request, "Ù…Ø¹Ø¯Ù† Ø§Ù†ØªØ®Ø§Ø¨â€ŒØ´Ø¯Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
            raise Exception("Ù…Ø¹Ø¯Ù† Ù†Ø§Ù…Ø¹ØªØ¨Ø±")

        # Ø°Ø®ÛŒØ±Ù‡ ÙˆØ²Ù†â€ŒÙ‡Ø§
        full_weight = request.POST.get('full_weight')
        empty_weight = request.POST.get('empty_weight')
        net_weight = request.POST.get('net_weight')

        data.pop('full_weight', None)
        data.pop('empty_weight', None)
        data.pop('net_weight', None)
        data.pop('mine_id', None)

        step = Step.objects.filter(order = 1).first()

        # Ø«Ø¨Øª Ù…ÙˆØ§Ø¯ Ø®Ø§Ù…
        for field, value in data.items():
            try:
                if value and float(value) > 0:
                    raw_material_obj = raw_material.objects.get(name=field)
                    coop_record = coops.objects.create(
                        user=request.user,
                        material=raw_material_obj,
                        quantity=Decimal(value),
                        state = step,
                        image=image
                    )
            except Exception as e:
                # raise Exception(f"Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øª Ù…ÙˆØ§Ø¯ Ø§ÙˆÙ„ÛŒÙ‡: {field}")
                print('Error is Save Raw amterial', field)

        if coop_record is None:
            raise Exception("Ù‡ÛŒÚ† Ú©ÙˆÙ¾ÛŒ Ø«Ø¨Øª Ù†Ø´Ø¯")

        # Ø«Ø¨Øª ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ø¯ÛŒÙ†Ø§Ù…ÛŒÚ©
        attributes = CoopAttribute.objects.filter(step=stepNumber)
        for attr in attributes:
            field_name = f'attr_{attr.id}'
            value = request.POST.get(field_name, '').strip()

            if attr.required and not value:
                messages.error(request, f'ÙÛŒÙ„Ø¯ "{attr.label}" Ø§Ù„Ø²Ø§Ù…ÛŒ Ø§Ø³Øª.', extra_tags='create_coop_error')
                raise Exception(f'ÙÛŒÙ„Ø¯ Ø§Ù„Ø²Ø§Ù…ÛŒ "{attr.label}" Ø®Ø§Ù„ÛŒ Ø§Ø³Øª')


            # âœ… Ø§Ú¯Ø± Ù†ÙˆØ¹ ÙÛŒÙ„Ø¯ date Ø¨ÙˆØ¯ØŒ ØªØ¨Ø¯ÛŒÙ„ Ø´Ù…Ø³ÛŒ Ø¨Ù‡ Ù…ÛŒÙ„Ø§Ø¯ÛŒ
            if attr.field_type == 'date':
                try:
                    # Ø§Ù†ØªØ¸Ø§Ø± Ø¯Ø§Ø±ÛŒÙ… ÙØ±Ù…Øª Ø¯Ø±ÛŒØ§ÙØªÛŒ Ø´Ù…Ø³ÛŒ Ù…Ø«Ù„ "1403/04/23" Ø¨Ø§Ø´Ø¯
                    jalali_parts = value.split('/')
                    if len(jalali_parts) == 3:
                        jy, jm, jd = map(int, jalali_parts)
                        jalali_date = jdatetime.date(jy, jm, jd)
                        gregorian_date = jalali_date.togregorian()
                        value = gregorian_date  # Ù†ÙˆØ¹: datetime.date
                except Exception as e:
                    print(f'Error converting jalali to gregorian for {field_name}: {value} => {e}')
                    continue  # Ø¯Ø± ØµÙˆØ±Øª Ø®Ø·Ø§ØŒ Ø§ÛŒÙ† Ù…Ù‚Ø¯Ø§Ø± Ø±Ø§ Ø°Ø®ÛŒØ±Ù‡ Ù†Ú©Ù†





            if value:
                CoopAttributeValue.objects.create(
                    coop=coop_record,
                    attribute=attr,
                    value=value,
                    user  = request.user
                )

        messages.success(request, "Ù…Ù‚Ø§Ø¯ÛŒØ± Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø«Ø¨Øª Ø´Ø¯Ù†Ø¯.")
        return render(request, 'success_page.html', {'content': 'Ø­ÙˆØ§Ù„Ù‡ Ø¬Ø¯ÛŒØ¯ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø«Ø¨Øª Ú¯Ø±Ø¯ÛŒØ¯'})

        # except Exception as e:
        #     print('âŒ Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øª Ø§Ø·Ù„Ø§Ø¹Ø§Øª:', e)

        #     # Ø§Ú¯Ø± Ú©ÙˆÙ¾ Ø³Ø§Ø®ØªÙ‡ Ø´Ø¯Ù‡ØŒ Ø¢Ù† Ø±Ø§ Ø­Ø°Ù Ú©Ù†
        #     if coop_record:
        #         coop_record.delete()

        #     messages.error(request, 'Ø®Ø·Ø§ Ø¯Ø± Ø°Ø®ÛŒØ±Ù‡ Ø§Ø·Ù„Ø§Ø¹Ø§Øª. ØªÙ…Ø§Ù… Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ Ø­Ø°Ù Ø´Ø¯Ù†Ø¯.')
        #     return redirect(request.path)

    else:
        # GET
        attributes = CoopAttribute.objects.filter(step=stepNumber)
        can_submit, is_confirmed = get_submit_and_confirmed(user=request.user, stepNumber=stepNumber)
        mother_materials = mother_material.objects.prefetch_related('mother_material').order_by('describe').all()
        mines = Mine.objects.all()


        steps = Step.objects.order_by('order')  # Ù…Ø±ØªØ¨â€ŒØ³Ø§Ø²ÛŒ Ù…Ø±Ø§Ø­Ù„


        return render(request, 'stone_section1.html', {
            'mother_materials': mother_materials,
            'is_confirmed': is_confirmed,
            'can_submit': can_submit,
            'mines': mines,
            'attributes': attributes,  # ğŸ‘ˆ Ø§Ø±Ø³Ø§Ù„ ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§ Ø¨Ù‡ Ù‚Ø§Ù„Ø¨
            # 'steps': steps,
        })




# def coop_state_detail(request, coop_id, state):

#     if state==STATE_CHOICES[0][0]:
#         ret = create_coope(request=request)
#         return ret

#     # coop = get_object_or_404(coops, id=coop_id)
#     # history = coop.state_history.filter(new_state=state).last()
#     # return render(request, 'coop/coop_state_detail.html', {
#     #     'coop': coop,
#     #     'state': state,
#     #     'history': history,
#     # })




from django.shortcuts import render, get_object_or_404, redirect
from .models import Step, coops, CoopStateHistory  # ÙØ±Ø¶ÛŒ
from django.http import Http404

def coop_state_detail(request, coop_id, state):
    try:
        step = Step.objects.get(url_name=state)
    except Step.DoesNotExist:
        return render(request, 'step_not_found.html', {
            'state': state,
            'coop_id': coop_id
        })

    if step.url_name == 'create_coope':
        return create_coope(request=request)

    coop = get_object_or_404(coops, id=coop_id)

    # ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ù…Ø±Ø¨ÙˆØ· Ø¨Ù‡ Ø§ÛŒÙ† Ù…Ø±Ø­Ù„Ù‡
    attributes = CoopAttribute.objects.filter(step=step)

    # Ù…Ù‚Ø§Ø¯ÛŒØ± Ø°Ø®ÛŒØ±Ù‡â€ŒØ´Ø¯Ù‡ Ù…Ø±Ø¨ÙˆØ· Ø¨Ù‡ Ø§ÛŒÙ† Ú©ÙˆÙ¾ Ùˆ Ø§ÛŒÙ† ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§
    attribute_values = {
        av.attribute_id: av.value
        for av in coop.attribute_values.filter(attribute__in=attributes)
    }

    # history = coop.state_history.filter(new_state=state).last()

    return render(request, 'coop_state_detail.html', {
        'coop': coop,
        'state': state,
        # 'history': history,
        'step': step,
        'attributes': attributes,
        'attribute_values': attribute_values
    })









def dynamic_step_view(request, url_name, order_id=None):
    try:

        if request.method == 'POST':
            coop_record = None  # Ø¯Ø± Ø³Ø·Ø­ Ø¨Ø§Ù„Ø§ ØªØ¹Ø±ÛŒÙ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ… Ú©Ù‡ Ø¯Ø± except Ù‡Ù… Ù‚Ø§Ø¨Ù„ Ø¯Ø³ØªØ±Ø³ÛŒ Ø¨Ø§Ø´Ø¯
            data = dict(request.POST.dict())
            data.pop('csrfmiddlewaretoken', None)

            image = None
            image_data = request.POST.get('image_data')
            if image_data:
                format, imgstr = image_data.split(';base64,')
                ext = format.split('/')[-1]
                image = ContentFile(base64.b64decode(imgstr), name='captured_image.' + ext)
                data.pop('image_data', None)


            step = Step.objects.get(url_name=url_name)
            stepNumber =step.order

            coop_record = coops.objects.filter(id=order_id).first()


            if coop_record is None:
                raise Exception("Ù‡ÛŒÚ† Ú©ÙˆÙ¾ÛŒ Ø«Ø¨Øª Ù†Ø´Ø¯")
            
            coop_record.state = step
            coop_record.save()


            # Ø«Ø¨Øª ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ø¯ÛŒÙ†Ø§Ù…ÛŒÚ©
            attributes = CoopAttribute.objects.filter(step=stepNumber)

            # Ø«Ø¨Øª Ù…ÙˆØ§Ø¯ Ø®Ø§Ù…
            for field, value in data.items():
                if not field.startswith('attr_'):
                    try:
                        if value and float(value) > 0:
                            raw_material_obj = raw_material.objects.get(name=field)
                            coop_record = coops.objects.create(
                                user=request.user,
                                material=raw_material_obj,
                                quantity=Decimal(value),
                                state = step,
                                image=image
                            )
                    except Exception as e:
                        # raise Exception(f"Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øª Ù…ÙˆØ§Ø¯ Ø§ÙˆÙ„ÛŒÙ‡: {field}")
                        print('Error is Save Raw amterial', field)




            for attr in attributes:
                field_name = f'attr_{attr.id}'




                # âœ³ï¸ Ø¨Ø®Ø´ Ø«Ø¨Øª CuttingSaw
                if attr.field_type == 'CuttingSaw':
                    lengths = request.POST.getlist('cutting_length[]')
                    widths = request.POST.getlist('cutting_width[]')
                    quantities = request.POST.getlist('cutting_quantity[]')
                    descriptions = request.POST.getlist('cutting_description[]')

                    for i in range(len(lengths)):
                        try:
                            length = float(lengths[i])
                            width = float(widths[i])
                            quantity = int(quantities[i])
                            description = descriptions[i]

                            if length > 0 and width > 0 and quantity > 0:
                                cutting_saw_obj  = CuttingSaw.objects.create(
                                    coop=coop_record,
                                    lenght=length,
                                    width=width,
                                    quantity=quantity,
                                    description=description
                                )

                                CoopAttributeValue.objects.create(
                                    coop=coop_record,
                                    attribute=attr,
                                    value=str(cutting_saw_obj.id),
                                    user  = request.user
                                )



                        except (ValueError, IndexError) as e:
                            print(f'Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øª CuttingSaw Ø¯Ø± Ø±Ø¯ÛŒÙ {i}: {e}')
                    continue  # Ú†ÙˆÙ† Ù…Ù‚Ø¯Ø§Ø± Ø¯Ø± Ù…Ø¯Ù„ Ø¯ÛŒÚ¯Ø± Ø°Ø®ÛŒØ±Ù‡ Ø´Ø¯ØŒ Ø§Ø¯Ø§Ù…Ù‡ Ù†Ø¯Ù‡



                # âœ³ï¸ Ø¨Ø®Ø´ Ø«Ø¨Øª CuttingAround
                elif attr.field_type == 'CuttingAround':
                    lengths = request.POST.getlist('cutting_around_length[]')
                    widths = request.POST.getlist('cutting_around_width[]')
                    quantities = request.POST.getlist('cutting_around_quantity[]')
                    serials = request.POST.getlist('cutting_around_serial[]')
                    descriptions = request.POST.getlist('cutting_around_description[]')

                    for i in range(len(lengths)):
                        try:
                            length = float(lengths[i])
                            width = float(widths[i])
                            quantity = int(quantities[i])
                            serial = int(serials[i])
                            description = descriptions[i]

                            if length > 0 and width > 0 and quantity > 0 and serial > 0:
                                cutting_around_obj = CuttingAround.objects.create(
                                    coop=coop_record,
                                    length=length,
                                    width=width,
                                    quantity=quantity,
                                    serial=serial,
                                    description=description
                                )

                                CoopAttributeValue.objects.create(
                                    coop=coop_record,
                                    attribute=attr,
                                    value=str(cutting_around_obj.id),
                                    user=request.user
                                )

                        except (ValueError, IndexError) as e:
                            print(f'Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øª CuttingAround Ø¯Ø± Ø±Ø¯ÛŒÙ {i}: {e}')
                    continue  # Ú†ÙˆÙ† Ù…Ù‚Ø¯Ø§Ø± Ø¯Ø± Ù…Ø¯Ù„ Ø¯ÛŒÚ¯Ø± Ø°Ø®ÛŒØ±Ù‡ Ø´Ø¯ØŒ Ø§Ø¯Ø§Ù…Ù‡ Ù†Ø¯Ù‡

                elif attr.field_type == 'multi_select':
                    values = request.POST.getlist(field_name)
                    if attr.required and not values:
                        messages.error(request, f'ÙÛŒÙ„Ø¯ "{attr.label}" Ø§Ù„Ø²Ø§Ù…ÛŒ Ø§Ø³Øª.', extra_tags='dynamic_coop_step_error')
                        raise Exception(f'ÙÛŒÙ„Ø¯ Ø§Ù„Ø²Ø§Ù…ÛŒ "{attr.label}" Ø®Ø§Ù„ÛŒ Ø§Ø³Øª')

                    # Ø­Ø°Ù Ù…Ù‚Ø§Ø¯ÛŒØ± Ù‚Ø¨Ù„ÛŒ
                    CoopAttributeValue.objects.filter(coop=coop_record, attribute=attr).delete()

                    # Ø°Ø®ÛŒØ±Ù‡ Ù…Ù‚Ø§Ø¯ÛŒØ± Ø¬Ø¯ÛŒØ¯
                    for val in values:
                        CoopAttributeValue.objects.create(
                            coop=coop_record,
                            attribute=attr,
                            value=val,
                            user  = request.user
                        )

                else:
                    value = request.POST.get(field_name, '').strip()


                    if attr.field_type == 'image':
                        value = request.FILES.get(field_name)

                    if attr.required and not value:
                        if attr.field_type !='bool':
                            messages.error(request, f'ÙÛŒÙ„Ø¯ "{attr.label}" Ø§Ù„Ø²Ø§Ù…ÛŒ Ø§Ø³Øª.', extra_tags='dynamic_coop_step_error')
                            raise Exception(f'ÙÛŒÙ„Ø¯ Ø§Ù„Ø²Ø§Ù…ÛŒ "{attr.label}" Ø®Ø§Ù„ÛŒ Ø§Ø³Øª')
                        
                   
                    CoopAttributeValue.objects.filter(coop=coop_record, attribute=attr).delete()


                    if value:
                        # Ø´Ø±Ø· Ø®Ø§Øµ Ø¨Ø±Ø§ÛŒ Ù…ÙˆØ§Ø¯ Ø§ÙˆÙ„ÛŒÙ‡: ÙÙ‚Ø· Ø§Ú¯Ø± Ù…Ù‚Ø¯Ø§Ø± Ø¹Ø¯Ø¯ÛŒ > 0 Ø¨ÙˆØ¯
                        if attr.field_type == 'material':
                            try:
                                if float(value) <= 0:
                                    continue  # Ø°Ø®ÛŒØ±Ù‡ Ù†Ú©Ù†
                            except ValueError:
                                continue  # Ø§Ú¯Ø± Ù…Ù‚Ø¯Ø§Ø± Ù†Ø§Ù…Ø¹ØªØ¨Ø± Ø¨ÙˆØ¯ Ù‡Ù… Ø°Ø®ÛŒØ±Ù‡ Ù†Ú©Ù†



                        # âœ… Ø§Ú¯Ø± Ù†ÙˆØ¹ ÙÛŒÙ„Ø¯ date Ø¨ÙˆØ¯ØŒ ØªØ¨Ø¯ÛŒÙ„ Ø´Ù…Ø³ÛŒ Ø¨Ù‡ Ù…ÛŒÙ„Ø§Ø¯ÛŒ
                        if attr.field_type == 'date':
                            try:
                                # Ø§Ù†ØªØ¸Ø§Ø± Ø¯Ø§Ø±ÛŒÙ… ÙØ±Ù…Øª Ø¯Ø±ÛŒØ§ÙØªÛŒ Ø´Ù…Ø³ÛŒ Ù…Ø«Ù„ "1403/04/23" Ø¨Ø§Ø´Ø¯
                                jalali_parts = value.split('/')
                                if len(jalali_parts) == 3:
                                    jy, jm, jd = map(int, jalali_parts)
                                    jalali_date = jdatetime.date(jy, jm, jd)
                                    gregorian_date = jalali_date.togregorian()
                                    value = gregorian_date  # Ù†ÙˆØ¹: datetime.date
                            except Exception as e:
                                print(f'Error converting jalali to gregorian for {field_name}: {value} => {e}')
                                continue  # Ø¯Ø± ØµÙˆØ±Øª Ø®Ø·Ø§ØŒ Ø§ÛŒÙ† Ù…Ù‚Ø¯Ø§Ø± Ø±Ø§ Ø°Ø®ÛŒØ±Ù‡ Ù†Ú©Ù†


                        CoopAttributeValue.objects.create(
                            coop=coop_record,
                            attribute=attr,
                            value=value,
                            user  = request.user
                        )







            messages.success(request, "Ù…Ù‚Ø§Ø¯ÛŒØ± Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø«Ø¨Øª Ø´Ø¯Ù†Ø¯.")
            return render(request, 'success_page.html', {'content': f'{step.title} Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø«Ø¨Øª Ú¯Ø±Ø¯ÛŒØ¯'})





        else:
            
            step = Step.objects.get(url_name=url_name)

            steps = Step.objects.order_by('order')  # Ù…Ø±ØªØ¨â€ŒØ³Ø§Ø²ÛŒ Ù…Ø±Ø§Ø­Ù„
            stepNumber =step.order
            attributes = CoopAttribute.objects.filter(step=step.order)

            mother_materials = mother_material.objects.prefetch_related('mother_material').order_by('describe').all()

            can_submit, is_confirmed = get_submit_and_confirmed(user=request.user, stepNumber=stepNumber)

            coop_record = coops.objects.filter(id=order_id).first()
            # Ø§Ú¯Ø± Ú©ÙˆÙ¾ ÙˆØ¬ÙˆØ¯ Ø¯Ø§Ø´ØªØŒ Ù…Ù‚Ø§Ø¯ÛŒØ± Ø«Ø¨Øª Ø´Ø¯Ù‡ Ù‚Ø¨Ù„ÛŒ Ø±Ø§ Ø¨Ú¯ÛŒØ±
            attribute_values = {}
            if coop_record:
                values = CoopAttributeValue.objects.filter(coop=coop_record)
                for val in values:
                    attribute_values[val.attribute.id] = val.value

            warehouses = Warehouse.objects.all()
            cutting_factories = Cutting_factory.objects.all()
            materials = raw_material.objects.all()
            cutting_saw_items = CuttingSaw.objects.filter(coop=coop_record)
            cutting_around_items = CuttingAround.objects.filter(coop=coop_record)

            # Ø§ÛŒÙ†Ø¬Ø§ Ù…ÛŒâ€ŒØªÙˆÙ†ÛŒ Ø¨Ø± Ø§Ø³Ø§Ø³ Ù†ÙˆØ¹ step Ø§Ù‚Ø¯Ø§Ù… Ø®Ø§ØµÛŒ Ø§Ù†Ø¬Ø§Ù… Ø¨Ø¯ÛŒ
            return render(request, 'step_placeholder.html', {
                'step': step,
                'order_id': order_id,
                'steps':steps,
                'is_confirmed': is_confirmed,
                'can_submit': can_submit,
                'mother_materials':mother_materials,
                'attributes':attributes,
                'attribute_values': attribute_values,  # ğŸ‘ˆ Ø§Ø±Ø³Ø§Ù„ Ø¨Ù‡ Ù‚Ø§Ù„Ø¨
                'warehouses': warehouses,
                'cutting_factories':cutting_factories,
                'materials': materials,
                'cutting_saw_items': cutting_saw_items,
                'cutting_around_items': cutting_around_items,
            })
    except Step.DoesNotExist:
        return render(request, 'step_not_found.html', {
            'url_name': url_name
        })









# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import AttributeGroupForm, CoopAttributeForm, CuttingFactoryForm, DriverRegisterForm, StepForm

from users.models import jobs,User,Profile
from django.db import transaction



def driver_list_view(request):
    drivers = Driver.objects.all()
    return render(request, 'drivers/driver_list.html', {'drivers': drivers})



def register_driver(request):
    if request.method == 'POST':
        form = DriverRegisterForm(request.POST)
        if form.is_valid():
            try:
                # Ø§ÛŒØ¬Ø§Ø¯ ØªØ±Ø§Ú©Ù†Ø´ Ø§Ù…Ù†
                with transaction.atomic():
                    # Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ø±Ø¨Ø±
                    user = User.objects.create_user(
                        username=form.cleaned_data['username'],
                        password=form.cleaned_data['password'],
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name'],
                    )


                    # Ø³Ø§Ø®Øª Ù¾Ø±ÙˆÙØ§ÛŒÙ„


                    # 1. Ú¯Ø±ÙØªÙ† Ù¾Ø±ÙˆÙØ§ÛŒÙ„ Ù…Ø±Ø¨ÙˆØ· Ø¨Ù‡ Ø¢Ù† Ú©Ø§Ø±Ø¨Ø±:
                    profile = get_object_or_404(Profile, user=user)

                    # 2. Ù…Ù‚Ø¯Ø§Ø± Ø¬Ø¯ÛŒØ¯ Ø¨Ø±Ø§ÛŒ job_position (Ù…Ø«Ù„Ø§Ù‹ Ø±Ø§Ù†Ù†Ø¯Ù‡):
                    driver_job = get_object_or_404(jobs, persian_name="Ø±Ø§Ù†Ù†Ø¯Ù‡")

                    # 3. Ø¢Ù¾Ø¯ÛŒØª Ù…Ù‚Ø¯Ø§Ø±
                    profile.job_position = driver_job
                    profile.first_name = form.cleaned_data['first_name']
                    profile.last_name = form.cleaned_data['last_name']

                    # 4. Ø°Ø®ÛŒØ±Ù‡ ØªØºÛŒÛŒØ±Ø§Øª
                    profile.save()




                    # Ø³Ø§Ø®Øª Ø´ÛŒØ¡ Ø±Ø§Ù†Ù†Ø¯Ù‡
                    driver = form.save(commit=False)
                    driver.user = user
                    driver.save()

                    messages.success(request, "Ø«Ø¨Øªâ€ŒÙ†Ø§Ù… Ø±Ø§Ù†Ù†Ø¯Ù‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯.")
                    return redirect('login')

            except Exception as e:
                # Ø¯Ø± ØµÙˆØ±Øª Ø®Ø·Ø§ Ø¯Ø± Ù‡Ø± Ù…Ø±Ø­Ù„Ù‡ØŒ Ú©Ø§Ø±Ø¨Ø± Ùˆ Ù¾Ø±ÙˆÙØ§ÛŒÙ„ Ø­Ø°Ù Ø´ÙˆØ¯
                if 'user' in locals():
                    user.delete()
                messages.error(request, f"Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øªâ€ŒÙ†Ø§Ù…: {str(e)}")
                return render(request, 'drivers/register_driver.html', {'form': form})

    else:
        form = DriverRegisterForm()
    
    return render(request, 'drivers/register_driver.html', {'form': form})



from .forms import DriverRegisterForm
from django.contrib import messages

def edit_driver(request, driver_id):
    driver = get_object_or_404(Driver, pk=driver_id)
    if request.method == 'POST':
        form = DriverRegisterForm(request.POST, instance=driver)
        if form.is_valid():
            form_instance = form.save(commit=False)
            # Ø¨Ù‡â€ŒØ±ÙˆØ²Ø±Ø³Ø§Ù†ÛŒ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ú©Ø§Ø±Ø¨Ø±
            user = driver.user
            user.username = form.cleaned_data['username']
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.save()

            form_instance.save()
            return redirect('driver_list')
    else:
        initial = {
            'username': driver.user.username,
            'first_name': driver.user.first_name,
            'last_name': driver.user.last_name,
        }
        form = DriverRegisterForm(instance=driver, initial=initial)

    return render(request, 'drivers/edit_driver.html', {'form': form})



def delete_driver_view(request, driver_id):
    driver = get_object_or_404(Driver, id=driver_id)

    if request.method == 'POST':
        user = driver.user
        driver.delete()
        user.delete()  # Ø­Ø°Ù ÛŒÙˆØ²Ø± Ù…Ø±ØªØ¨Ø·
        messages.success(request, "Ø±Ø§Ù†Ù†Ø¯Ù‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø­Ø°Ù Ø´Ø¯.")
        return redirect('driver_list')

    return render(request, 'drivers/delete_confirm.html', {'driver': driver})





from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.urls import reverse

@login_required
def manage_coop_attributes(request):
    form = CoopAttributeForm()
    attributes = CoopAttribute.objects.all()

    if request.method == 'POST':
        if 'edit_id' in request.POST:
            attr = get_object_or_404(CoopAttribute, id=request.POST['edit_id'])
            form = CoopAttributeForm(request.POST, instance=attr)
            if form.is_valid():
                form.save()
                messages.success(request, 'ÙˆÛŒÚ˜Ú¯ÛŒ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª ÙˆÛŒØ±Ø§ÛŒØ´ Ø´Ø¯.', extra_tags='create_coop_feature_success')
                return redirect('manage_coop_attributes')
        elif 'delete_id' in request.POST:
            attr = get_object_or_404(CoopAttribute, id=request.POST['delete_id'])
            attr.delete()
            messages.success(request, 'ÙˆÛŒÚ˜Ú¯ÛŒ Ø­Ø°Ù Ø´Ø¯.', extra_tags='create_coop_feature_success')
            return redirect('manage_coop_attributes')
        else:
            form = CoopAttributeForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'ÙˆÛŒÚ˜Ú¯ÛŒ Ø¬Ø¯ÛŒØ¯ Ø§Ø¶Ø§ÙÙ‡ Ø´Ø¯.', extra_tags='create_coop_feature_success')
                return redirect('manage_coop_attributes')
            else:
                # Print errors to console (terminal) for debugging
                print(form.errors)
                messages.error(request, ' Ù†Ø§Ù… ÙˆÛŒÚ˜Ú¯ÛŒ ØªÚ©Ø±Ø§Ø±ÛŒ Ø§Ø³Øª ÛŒØ§ ÙØ±Ù… Ú©Ø§Ù…Ù„ Ù¾Ø±Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.', extra_tags='create_coop_feature_error')

    return render(request, 'manage_attributes.html', {
        'form': form,
        'attributes': attributes
    })





from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from .models import Step, StepAccess

def manage_access(request):
    # users = User.objects.all()
    # steps = Step.objects.all()

    query = request.GET.get('q', '')
    if query:
        users = User.objects.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(username__icontains=query)
        )
    else:
        users = User.objects.all()

    steps = Step.objects.all()



    if request.method == 'POST':
        # Receive data from form: 
        # expect keys like access_USERID_STEPID = 'view' or 'submit' or ''(no access)
        for user in users:
            for step in steps:
                key = f'access_{user.id}_{step.id}'
                level = request.POST.get(key, '')
                # Update or delete StepAccess accordingly
                if level in ['view', 'submit']:
                    obj, created = StepAccess.objects.update_or_create(
                        user=user, step=step,
                        defaults={'access_level': level}
                    )
                else:
                    StepAccess.objects.filter(user=user, step=step).delete()
        return redirect('manage_access')  # redirect to refresh page

    # Prepare current access matrix
    access_matrix = {}
    for user in users:
        access_matrix[user.id] = {}
        for step in steps:
            try:
                sa = StepAccess.objects.get(user=user, step=step)
                access_matrix[user.id][step.id] = sa.access_level
            except StepAccess.DoesNotExist:
                access_matrix[user.id][step.id] = ''

    context = {
        'users': users,
        'steps': steps,
        'access_matrix': access_matrix,
    }
    return render(request, 'manage_access.html', context)





def steps_list(request):
    steps = Step.objects.order_by('order')
    return render(request, 'steps_list.html', {'steps': steps})


def manage_step(request, step_id=None):
    if step_id:
        step = get_object_or_404(Step, id=step_id)
    else:
        step = None

    form = StepForm(request.POST or None, instance=step)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('steps_list')  # Ø¢Ø¯Ø±Ø³ Ù„ÛŒØ³Øª Ù…Ø±Ø§Ø­Ù„

    return render(request, 'manage_step.html', {
        'form': form,
        'is_edit': step is not None
    })



def delete_step(request, step_id):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Ø§Ø¬Ø§Ø²Ù‡ Ø¯Ø³ØªØ±Ø³ÛŒ Ù†Ø¯Ø§Ø±ÛŒØ¯.")

    step = get_object_or_404(Step, id=step_id)
    step.delete()
    messages.success(request, "Ù…Ø±Ø­Ù„Ù‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø­Ø°Ù Ø´Ø¯.")
    return redirect('steps_list')






from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import AttributeGroup, CoopAttribute
from .forms import AttributeGroupForm

def manage_attribute_groups(request):
    groups = AttributeGroup.objects.prefetch_related('attributes')
    form = AttributeGroupForm()

    edit_group_obj = None
    edit_form = None

    if request.method == 'POST':
        edit_id = request.POST.get('edit_id')
        if edit_id:
            # ÙˆÛŒØ±Ø§ÛŒØ´
            edit_group_obj = get_object_or_404(AttributeGroup, pk=edit_id)
            edit_form = AttributeGroupForm(request.POST, instance=edit_group_obj)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, 'âœ… Ú¯Ø±ÙˆÙ‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª ÙˆÛŒØ±Ø§ÛŒØ´ Ø´Ø¯.')
                return redirect('attribute_group_view')
        else:
            # Ø³Ø§Ø®Øª Ø¬Ø¯ÛŒØ¯
            form = AttributeGroupForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'âœ… Ú¯Ø±ÙˆÙ‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø³Ø§Ø®ØªÙ‡ Ø´Ø¯.')
                return redirect('attribute_group_view')

    elif request.method == 'GET' and 'edit_id' in request.GET:
        edit_id = request.GET['edit_id']
        edit_group_obj = get_object_or_404(AttributeGroup, pk=edit_id)
        edit_form = AttributeGroupForm(instance=edit_group_obj)

    return render(request, 'attribute_group.html', {
        'form': form,
        'groups': groups,
        'edit_group_obj': edit_group_obj,
        'edit_form': edit_form
    })

def delete_group(request, pk):
    group = get_object_or_404(AttributeGroup, pk=pk)
    group.delete()
    messages.success(request, 'ğŸ—‘ï¸ Ú¯Ø±ÙˆÙ‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø­Ø°Ù Ø´Ø¯.')
    return redirect('attribute_group_view')  # Ù…Ø³ÛŒØ± Ù…Ù†Ø§Ø³Ø¨ Ø±Ø§ Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡





@login_required
def export_group_excel(request, group_id):
    group = get_object_or_404(AttributeGroup, id=group_id)
    coop_values = CoopAttributeValue.objects.filter(attribute__in=group.attributes.all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Ú¯Ø²Ø§Ø±Ø´ ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§"

    ws.append(["ÙˆÛŒÚ˜Ú¯ÛŒ", "Ù…Ù‚Ø¯Ø§Ø±", "Ù…Ø±Ø¨ÙˆØ· Ø¨Ù‡ Ú©ÙˆÙ¾", "ØªØ§Ø±ÛŒØ®"])

    for value in coop_values:
        ws.append([
            value.attribute.label,
            value.value,
            value.coop.id if value.coop else '',
            value.created_at.strftime("%Y-%m-%d") if hasattr(value, 'created_at') else ''
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{group.name}.xlsx"'
    wb.save(response)
    return response



import pythoncom
import win32com.client

def convert_excel_to_pdf(excel_path, pdf_path):
    pythoncom.CoInitialize()  # ğŸ‘ˆ Ù…Ù‡Ù…
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False

        wb = excel.Workbooks.Open(excel_path)
        wb.ExportAsFixedFormat(0, pdf_path)  # 0 = PDF
        wb.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()


def convert_str_price2float(price:str):

        if price is not None:
            if price !='None':

                price = float(price.replace(',', ''))
                return price
        
        return 0


@login_required
def create_preinvoice_view(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        customer = Buyer.objects.get(id=customer_id)

        selected_ids = request.POST.getlist('selected_coops')
        selected_coops = coops.objects.filter(id__in=selected_ids)

        # Load template




        # ØªØ§Ø±ÛŒØ® Ø§Ù…Ø±ÙˆØ² Ù…ÛŒÙ„Ø§Ø¯ÛŒ
        today_gregorian = datetime.today()




        language = request.POST.get('language', 'fa')
        # Retrieve other form data like customer, coops, prices...

        if language == 'fa':
            template_path = 'media/templates/preinvoice_template.xlsx'
            # ØªØ¨Ø¯ÛŒÙ„ Ø¨Ù‡ Ø¬Ù„Ø§Ù„ÛŒ
            today_gregorian = jdatetime.datetime.fromgregorian(datetime=today_gregorian)


        else:
            template_path = 'media/templates/en_preinvoice_template.xlsx'

        today_gregorian = today_gregorian.strftime('%Y/%m/%d')


        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Write customer info
        ws['B5'] =  f'{customer.first_name} {customer.last_name}'
        ws['G5'] =  f'{customer.phone_number}'
        ws['B6'] =  f'{customer.address}'
        ws['G4'] =  f'{today_gregorian}'

        # Write coop rows (starting from row 5)
        row =8
        col_start = 1

        total_price = 0
        total_discount = 0

        for iter,coop in enumerate(selected_coops):
            material_name = coop.material.name
            quantity = coop.quantity
            # try:
            price = (request.POST.get(f'price_{coop.id}', '0'))
            price = convert_str_price2float(price=price)
            # except:
            #     price = 0
            discount = float(request.POST.get(f'discount_{coop.id}', '0'))
            total = quantity * price * (100 - discount) / 100

            ws.cell(row=row, column=col_start, value=iter+1)
            ws.cell(row=row, column=col_start+1, value=material_name)
            ws.cell(row=row, column=col_start+2, value=0)
            ws.cell(row=row, column=col_start+3, value=quantity)
            ws.cell(row=row, column=col_start+4, value=float(price))
            ws.cell(row=row, column=col_start+5, value=float(discount))
            ws.cell(row=row, column=col_start+6, value=float(total))
            total_price+=(float(total))
            total_discount+=(float(discount))
            row += 1

        ws['G15'] =  total_price
        ws['G16'] =  total_discount
        ws['G17'] =  total_price - total_discount

        # Save to temp file and return as response
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            file_data = tmp.read()

        filename_base = f"preinvoice_{customer.first_name}_{customer.last_name}_{uuid.uuid4().hex[:6]}"
        media_dir = os.path.join(settings.MEDIA_ROOT, "preinvoices")
        os.makedirs(media_dir, exist_ok=True)

        excel_path = os.path.join(media_dir, f"{filename_base}.xlsx")
        pdf_path = os.path.join(media_dir, f"{filename_base}.pdf")

        # Save Excel file
        wb.save(excel_path)

        # Convert to PDF
        convert_excel_to_pdf(excel_path, pdf_path)

        # Redirect to preview page
        return redirect("preview_preinvoice", filename=filename_base)



        response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="preinvoice_{customer.first_name}_{customer.last_name}_{time.time()}.xlsx"'
        os.unlink(tmp.name)  # Delete the temp file
        return response

    else:
        coops_final = coops.objects.filter(state__order=Step.objects.latest('order').order)

        for coop in coops_final:
            coop.total_price =  calculate_total_price(coop=coop)
            try:
                sell_price = coop.attribute_values.filter(attribute__label="Ù‚ÛŒÙ…Øª ÙØ±ÙˆØ´").first()
            
                # print(coop.sell)
                if sell_price is not None:
                    if sell_price.value !='None':
                        sell_price = sell_price.value
                        sell_price = (sell_price.replace(':', ''))
                        sell_price = (sell_price.replace(' ', ''))
                        # sell_price = Decimal(number)
                    else:
                        sell_price =0 
                else:
                    sell_price = 0

            except:
                sell_price = 0


            coop.sell_price = sell_price

        customers = Buyer.objects.all()

        return render(request, 'create_preinvoice.html', {
            'coops': coops_final,
            'customers': customers,
        })
    

def show_preinvoce_result(request,filename):

    context = {
        
        'pdf_url': request.build_absolute_uri(settings.MEDIA_URL + f"preinvoices/{filename}.pdf"),
        'excel_url': request.build_absolute_uri(settings.MEDIA_URL + f"preinvoices/{filename}.xlsx"),
        'image_url': request.build_absolute_uri(settings.MEDIA_URL + f"preinvoices/{filename}.png"),  # Optional
    }

    return render(request, 'preinvoice_result.html', context)



@csrf_exempt  # optional, only if CSRF becomes a problem with external posts
def create_english_invoice(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        if customer_id is None:
            return HttpResponseForbidden("Ù…Ø´ØªØ±ÛŒ ÛŒØ§ÙØª Ù†Ø´Ø¯")

        customer = Buyer.objects.get(id=customer_id)

        selected_ids = request.POST.getlist('selected_coops')
        selected_coops = coops.objects.filter(id__in=selected_ids)

        # Load template



        template_path = 'media/templates/en_preinvoice_template.xlsx'

        today_gregorian = today_gregorian.strftime('%Y/%m/%d')


        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Write customer info
        ws['B5'] =  f'{customer.first_name} {customer.last_name}'
        ws['G5'] =  f'{customer.phone_number}'
        ws['B6'] =  f'{customer.address}'

        # Write coop rows (starting from row 5)
        row =8
        col_start = 1

        total_price = 0
        total_discount = 0

        for iter,coop in enumerate(selected_coops):
            material_name = coop.material.name
            quantity = coop.quantity
            try:
                price = float(request.POST.get(f'price_{coop.id}', '0'))
            except:
                price = 0
            discount = float(request.POST.get(f'discount_{coop.id}', '0'))
            total = quantity * price * (100 - discount) / 100

            ws.cell(row=row, column=col_start, value=iter+1)
            ws.cell(row=row, column=col_start+1, value=material_name)
            ws.cell(row=row, column=col_start+2, value=0)
            ws.cell(row=row, column=col_start+3, value=quantity)
            ws.cell(row=row, column=col_start+4, value=float(price))
            ws.cell(row=row, column=col_start+5, value=float(discount))
            ws.cell(row=row, column=col_start+6, value=float(total))
            total_price+=(float(total))
            total_discount+=(float(discount))
            row += 1

        ws['G15'] =  total_price
        ws['G16'] =  total_discount
        ws['G17'] =  total_price - total_discount

        # Save to temp file and return as response
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            file_data = tmp.read()

        filename_base = f"preinvoice_{customer.first_name}_{customer.last_name}_{uuid.uuid4().hex[:6]}"
        media_dir = os.path.join(settings.MEDIA_ROOT, "preinvoices")
        os.makedirs(media_dir, exist_ok=True)

        excel_path = os.path.join(media_dir, f"{filename_base}.xlsx")
        # pdf_path = os.path.join(media_dir, f"{filename_base}.pdf")

        # Save to in-memory file
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Return response
        response = HttpResponse(
            file_stream,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={excel_path}'
        return response

    return HttpResponse("Only POST allowed", status=405)






@login_required
def cutting_factory_view(request):
    jobs_qs = Cutting_factory.objects.all()
    return render(request, 'model_template/list.html', {'jobs': jobs_qs})

@login_required
def cutting_factory_create_view(request):
    if request.method == 'POST':
        form = CuttingFactoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ú©Ø§Ø±Ø®Ø§Ù†Ù‡ Ø¬Ø¯ÛŒØ¯ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø§ÛŒØ¬Ø§Ø¯ Ø´Ø¯.')
            return redirect('cutting_factory_list')
        
    else:
        form = CuttingFactoryForm()
    return render(request, 'model_template/form.html', {'form': form, 'title': 'Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ø±Ø®Ø§Ù†Ù‡ Ø¬Ø¯ÛŒØ¯'})

@login_required
def cutting_factory_edit_view(request, pk):
    factory = get_object_or_404(Cutting_factory, pk=pk)
    if request.method == 'POST':
        form = CuttingFactoryForm(request.POST, instance=factory)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ø´ØºÙ„ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª ÙˆÛŒØ±Ø§ÛŒØ´ Ø´Ø¯.')
            return redirect('cutting_factory_list')
    else:
        form = CuttingFactoryForm(instance=factory)
    return render(request, 'model_template/form.html', {'form': form, 'title': 'ÙˆÛŒØ±Ø§ÛŒØ´ Ú©Ø§Ø±Ø®Ø§Ù†Ù‡'})

@login_required
def cutting_factory_delete_view(request, pk):
    job = get_object_or_404(Cutting_factory, pk=pk)
    if request.method == 'POST':
        job.delete()
        messages.success(request, 'Ú©Ø§Ø±Ø®Ø§Ù†Ù‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø­Ø°Ù Ø´Ø¯.')
        return redirect('cutting_factory_list')
    return render(request, 'model_template/confirm_delete.html', {'job': job})




def calculate_total_price(coop):
    total = Decimal(0)
    # Get all price attributes with multipliers
    price_attrs = PriceAttribute.objects.select_related('attribute').all()

    for price_attr in price_attrs:
        attr = price_attr.attribute
        # Get the CoopAttributeValue for this coop and attribute
        try:
            attr_value_obj = CoopAttributeValue.objects.get(coop=coop, attribute=attr)
            if attr_value_obj.value is not None:
                if attr_value_obj.value !='None':

                    number = float(attr_value_obj.value.replace(',', ''))
                    attr_value = Decimal(number)
                    total += attr_value * price_attr.multiplier
        except CoopAttributeValue.DoesNotExist:
            continue  # Attribute value not set for this coop, skip

    total =  f"{total:,.0f}"  # Ø¨Ø¯ÙˆÙ† Ø±Ù‚Ù… Ø§Ø¹Ø´Ø§Ø±ØŒ Ø¬Ø¯Ø§Ú©Ù†Ù†Ø¯Ù‡ Ù‡Ø²Ø§Ø±Ú¯Ø§Ù† Ú©Ø§Ù…Ø§
    return total




def price_attribute_list(request):
    # Ú¯Ø±ÙØªÙ† ØªÙ…Ø§Ù… CoopAttribute Ù‡Ø§ÛŒ Ù†ÙˆØ¹ 'price'
    price_attrs = CoopAttribute.objects.filter(field_type='price').order_by('label')

    # Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² ÙˆØ¬ÙˆØ¯ PriceAttribute Ø¨Ø±Ø§ÛŒ Ù‡Ù…Ù‡â€ŒÛŒ Ø§ÛŒÙ†Ù‡Ø§ØŒ Ø¯Ø± ØµÙˆØ±Øª Ù†Ø¨ÙˆØ¯ Ø¨Ø³Ø§Ø²
    for attr in price_attrs:
        PriceAttribute.objects.get_or_create(attribute=attr)

    # Ú¯Ø±ÙØªÙ† PriceAttributeÙ‡Ø§ÛŒ Ù…Ø±Ø¨ÙˆØ·Ù‡
    price_attrs_with_multiplier = PriceAttribute.objects.filter(attribute__in=price_attrs).select_related('attribute')

    if request.method == 'POST':
        # Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ±Ù…
        for pa in price_attrs_with_multiplier:
            key = f'multiplier_{pa.attribute.id}'
            multiplier_str = request.POST.get(key)
            if multiplier_str:
                try:
                    pa.multiplier = float(multiplier_str)
                    pa.save()
                except ValueError:
                    messages.error(request, f'Ø¶Ø±ÛŒØ¨ Ø¨Ø±Ø§ÛŒ "{pa.attribute.label}" Ø¹Ø¯Ø¯ Ù…Ø¹ØªØ¨Ø± Ù†ÛŒØ³Øª.')
                    return redirect('price_attribute_list')

        messages.success(request, 'Ø¶Ø±ÛŒØ¨â€ŒÙ‡Ø§ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø°Ø®ÛŒØ±Ù‡ Ø´Ø¯Ù†Ø¯.')
        return redirect('price_attribute_list')

    context = {
        'price_attrs': price_attrs_with_multiplier,
    }
    return render(request, 'price/price_attribute_list.html', context)


from django.core.exceptions import PermissionDenied

def is_level_one(user):
    return hasattr(user, 'profile') and user.profile.job_position.level <= 1



@login_required
def request_delete_coop(request, coop_id):
    coop = get_object_or_404(coops, id=coop_id, is_active=True)

    # Prevent duplicates
    if CoopDeleteRequest.objects.filter(coop=coop, approved=None).exists():
        messages.warning(request, "Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø­Ø°Ù Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† Ø¨Ø§Ø± Ù‚Ø¨Ù„Ø§Ù‹ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.")
        return redirect('coop_list')

    CoopDeleteRequest.objects.create(
        coop=coop,
        requested_by=request.user
    )
    messages.success(request, "Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø­Ø°Ù Ø«Ø¨Øª Ø´Ø¯ Ùˆ Ø¯Ø± Ø§Ù†ØªØ¸Ø§Ø± ØªØ£ÛŒÛŒØ¯ Ù…Ø¯ÛŒØ± Ø§Ø³Øª.")
    return redirect('coop_list')

@login_required
@user_passes_test(is_level_one)
def review_delete_requests(request):
    requests = CoopDeleteRequest.objects.filter(approved=None).select_related('coop', 'requested_by')
    return render(request, 'delete_requests.html', {'requests': requests})



@login_required
@user_passes_test(is_level_one)
def approve_delete_request(request, coop_id):
    from django.utils import timezone

    delete_request = get_object_or_404(CoopDeleteRequest, id=coop_id, approved=None)
    delete_request.approved = True
    delete_request.reviewed_by = request.user
    delete_request.reviewed_at = timezone.now()
    delete_request.save()

    # Soft delete the coop
    coop = delete_request.coop
    coop.is_active = False
    coop.set_changed_by(request.user)
    coop.save()

    messages.success(request, f"Ø¨Ø§Ø± #{coop.id} Ø­Ø°Ù Ø´Ø¯.")
    return redirect('review_delete_requests')



@login_required
@user_passes_test(is_level_one)
def reject_delete_request(request, coop_id):
    from django.utils import timezone

    delete_request = get_object_or_404(CoopDeleteRequest, id=coop_id, approved=None)
    delete_request.approved = False
    delete_request.reviewed_by = request.user
    delete_request.reviewed_at = timezone.now()
    delete_request.comment = request.POST.get("comment", "")
    delete_request.save()

    messages.info(request, "Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø­Ø°Ù Ø±Ø¯ Ø´Ø¯.")
    return redirect('review_delete_requests')

 

